import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Productos para Importar"

# Headers matching InventoryImportRow interface
headers = [
    "code", "barcode", "name", "description", "category", "categoryColor",
    "price", "cost", "stock", "minStock", "criticalStock", "expiryDate", "location"
]

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill(start_color='465fff', end_color='465fff', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 100 products grouped by category
products = [
    # === ANTIBIÓTICOS (10) ===
    ("AMX500", "7701000000001", "Amoxicilina 500 mg", "Caja x 21 cápsulas", "Antibióticos", "#DC2626", 18500, 12000, 45, 10, 5, "2027-06-15", "Estante A-1"),
    ("AMP500", "7701000000002", "Ampicilina 500 mg", "Caja x 24 cápsulas", "Antibióticos", "#DC2626", 22000, 14500, 30, 8, 4, "2027-03-20", "Estante A-1"),
    ("AZI500", "7701000000003", "Azitromicina 500 mg", "Caja x 6 tabletas", "Antibióticos", "#DC2626", 35000, 23000, 25, 5, 3, "2027-09-10", "Estante A-2"),
    ("CEF500", "7701000000004", "Cefalexina 500 mg", "Caja x 24 cápsulas", "Antibióticos", "#DC2626", 15000, 9500, 50, 10, 5, "2027-12-01", "Estante A-2"),
    ("CIP500", "7701000000005", "Ciprofloxacino 500 mg", "Caja x 12 tabletas", "Antibióticos", "#DC2626", 12000, 7500, 40, 10, 5, "2027-08-15", "Estante A-3"),
    ("CLX300", "7701000000006", "Clindamicina 300 mg", "Caja x 16 cápsulas", "Antibióticos", "#DC2626", 28000, 18000, 20, 5, 3, "2027-05-20", "Estante A-3"),
    ("DOX100", "7701000000007", "Doxiciclina 100 mg", "Caja x 10 cápsulas", "Antibióticos", "#DC2626", 16000, 10000, 35, 8, 4, "2027-11-30", "Estante A-4"),
    ("ERY500", "7701000000008", "Eritromicina 500 mg", "Caja x 24 tabletas", "Antibióticos", "#DC2626", 14000, 9000, 28, 8, 4, "2027-07-25", "Estante A-4"),
    ("LEV500", "7701000000009", "Levofloxacino 500 mg", "Caja x 10 tabletas", "Antibióticos", "#DC2626", 32000, 21000, 18, 5, 3, "2027-10-15", "Estante A-5"),
    ("MET500", "7701000000010", "Metronidazol 500 mg", "Caja x 24 tabletas", "Antibióticos", "#DC2626", 8500, 5000, 60, 15, 8, "2028-01-10", "Estante A-5"),

    # === ANALGÉSICOS (10) ===
    ("IBU400", "7701000000011", "Ibuprofeno 400 mg", "Caja x 20 tabletas", "Analgésicos", "#2563EB", 12500, 8300, 80, 20, 10, "2027-12-15", "Estante B-1"),
    ("ACE500", "7701000000012", "Acetaminofén 500 mg", "Caja x 100 tabletas", "Analgésicos", "#2563EB", 9800, 6100, 100, 25, 12, "2028-03-20", "Estante B-1"),
    ("DIC50", "7701000000013", "Diclofenaco 50 mg", "Caja x 20 tabletas", "Analgésicos", "#2563EB", 11000, 7000, 55, 12, 6, "2027-09-30", "Estante B-2"),
    ("NAP500", "7701000000014", "Naproxeno 500 mg", "Caja x 20 tabletas", "Analgésicos", "#2563EB", 15500, 10000, 40, 10, 5, "2027-11-15", "Estante B-2"),
    ("KET10", "7701000000015", "Ketorolaco 10 mg", "Caja x 10 tabletas", "Analgésicos", "#2563EB", 8000, 5000, 70, 15, 8, "2027-08-25", "Estante B-3"),
    ("IND25", "7701000000016", "Indometacina 25 mg", "Caja x 30 cápsulas", "Analgésicos", "#2563EB", 13000, 8500, 30, 8, 4, "2027-06-10", "Estante B-3"),
    ("PIR20", "7701000000017", "Piroxicam 20 mg", "Caja x 30 cápsulas", "Analgésicos", "#2563EB", 17000, 11000, 25, 6, 3, "2027-10-20", "Estante B-4"),
    ("TRM50", "7701000000018", "Tramadol 50 mg", "Caja x 20 tabletas", "Analgésicos", "#2563EB", 25000, 16000, 20, 5, 3, "2027-07-15", "Estante B-4"),
    ("MEL15", "7701000000019", "Meloxicam 15 mg", "Caja x 10 tabletas", "Analgésicos", "#2563EB", 14500, 9500, 35, 8, 4, "2028-02-28", "Estante B-5"),
    ("CEX200", "7701000000020", "Celecoxib 200 mg", "Caja x 20 cápsulas", "Analgésicos", "#2563EB", 45000, 30000, 15, 5, 2, "2028-01-15", "Estante B-5"),

    # === VITAMINAS (10) ===
    ("VTC1G", "7701000000021", "Vitamina C 1000 mg", "Frasco x 30 tabletas", "Vitaminas", "#16A34A", 21400, 15400, 50, 15, 8, "2028-06-10", "Estante C-1"),
    ("VTD1K", "7701000000022", "Vitamina D3 1000 UI", "Frasco x 60 cápsulas", "Vitaminas", "#16A34A", 28000, 18000, 40, 10, 5, "2028-08-15", "Estante C-1"),
    ("VTE400", "7701000000023", "Vitamina E 400 UI", "Frasco x 50 cápsulas", "Vitaminas", "#16A34A", 32000, 21000, 35, 8, 4, "2028-04-20", "Estante C-2"),
    ("VTB12", "7701000000024", "Vitamina B12 500 mcg", "Frasco x 60 tabletas", "Vitaminas", "#16A34A", 18500, 12000, 45, 12, 6, "2028-03-10", "Estante C-2"),
    ("ACF400", "7701000000025", "Ácido Fólico 400 mcg", "Caja x 30 tabletas", "Vitaminas", "#16A34A", 12000, 7500, 60, 15, 8, "2028-07-25", "Estante C-3"),
    ("ZNC50", "7701000000026", "Zinc 50 mg", "Frasco x 60 tabletas", "Vitaminas", "#16A34A", 15000, 9500, 55, 12, 6, "2028-05-15", "Estante C-3"),
    ("HRO50", "7701000000027", "Hierro 50 mg", "Caja x 30 tabletas", "Vitaminas", "#16A34A", 9500, 6000, 70, 18, 9, "2028-02-28", "Estante C-4"),
    ("CAL600", "7701000000028", "Calcio 600 mg + Vit D", "Frasco x 60 tabletas", "Vitaminas", "#16A34A", 24000, 15500, 38, 10, 5, "2028-09-10", "Estante C-4"),
    ("OMG3", "7701000000029", "Omega 3 1000 mg", "Frasco x 30 cápsulas blandas", "Vitaminas", "#16A34A", 38000, 25000, 28, 8, 4, "2028-01-20", "Estante C-5"),
    ("MULTI", "7701000000030", "Multivitamínico Completo", "Frasco x 60 tabletas", "Vitaminas", "#16A34A", 42000, 28000, 32, 8, 4, "2028-11-15", "Estante C-5"),

    # === ANTIGRIPALES (8) ===
    ("GRIPAL", "7701000000031", "Antigripal Day", "Caja x 12 tabletas", "Antigripales", "#F97316", 6500, 3800, 90, 20, 10, "2028-04-15", "Estante D-1"),
    ("CLF300", "7701000000032", "Clorfenamina 4 mg", "Caja x 20 tabletas", "Antigripales", "#F97316", 5500, 3200, 60, 15, 8, "2027-12-30", "Estante D-1"),
    ("PSE60", "7701000000033", "Pseudoefedrina 60 mg", "Caja x 12 tabletas", "Antigripales", "#F97316", 11000, 7000, 40, 10, 5, "2027-09-20", "Estante D-2"),
    ("DEX15", "7701000000034", "Dextrometorfano 15 mg", "Jarabe 120 ml", "Antigripales", "#F97316", 12500, 8000, 50, 12, 6, "2027-11-10", "Estante D-2"),
    ("GUA100", "7701000000035", "Guaifenesina 100 mg", "Jarabe 150 ml", "Antigripales", "#F97316", 9800, 6000, 45, 10, 5, "2027-08-05", "Estante D-3"),
    ("LOR10", "7701000000036", "Loratadina 10 mg", "Caja x 20 tabletas", "Antigripales", "#F97316", 7500, 4500, 70, 15, 8, "2028-03-15", "Estante D-3"),
    ("CET10", "7701000000037", "Cetirizina 10 mg", "Caja x 20 tabletas", "Antigripales", "#F97316", 8000, 5000, 55, 12, 6, "2028-01-25", "Estante D-4"),
    ("FLU50", "7701000000038", "Fluimucil 600 mg", "Caja x 10 sobres", "Antigripales", "#F97316", 18000, 12000, 35, 8, 4, "2027-10-30", "Estante D-4"),

    # === GASTROINTESTINALES (8) ===
    ("OMP20", "7701000000039", "Omeprazol 20 mg", "Caja x 30 cápsulas", "Gastrointestinales", "#8B5CF6", 14000, 9000, 60, 15, 8, "2028-05-10", "Estante E-1"),
    ("LNS30", "7701000000040", "Lansoprazol 30 mg", "Caja x 30 cápsulas", "Gastrointestinales", "#8B5CF6", 22000, 14500, 35, 8, 4, "2028-02-20", "Estante E-1"),
    ("RNT150", "7701000000041", "Ranitidina 150 mg", "Caja x 20 tabletas", "Gastrointestinales", "#8B5CF6", 9500, 5800, 50, 12, 6, "2027-11-15", "Estante E-2"),
    ("LOP2", "7701000000042", "Loperamida 2 mg", "Caja x 24 cápsulas", "Gastrointestinales", "#8B5CF6", 6500, 3800, 80, 20, 10, "2028-04-30", "Estante E-2"),
    ("SMZ500", "7701000000043", "Sales de Rehidratación", "Sobre 30 g", "Gastrointestinales", "#8B5CF6", 3500, 2000, 100, 25, 12, "2029-01-15", "Estante E-3"),
    ("DIM20", "7701000000044", "Dimeticona 40 mg", "Caja x 30 tabletas masticables", "Gastrointestinales", "#8B5CF6", 11000, 7000, 40, 10, 5, "2027-09-25", "Estante E-3"),
    ("METOC", "7701000000045", "Metoclopramida 10 mg", "Caja x 30 tabletas", "Gastrointestinales", "#8B5CF6", 7500, 4500, 55, 12, 6, "2027-12-10", "Estante E-4"),
    ("PNC4", "7701000000046", "Probióticos 4 cepas", "Frasco x 20 cápsulas", "Gastrointestinales", "#8B5CF6", 28000, 18000, 25, 6, 3, "2027-08-20", "Estante E-4"),

    # === CARDIOVASCULARES (8) ===
    ("LOS50", "7701000000047", "Losartán 50 mg", "Caja x 30 tabletas", "Cardiovasculares", "#DC2626", 18000, 12000, 50, 12, 6, "2028-06-15", "Estante F-1"),
    ("ATN10", "7701000000048", "Atenolol 50 mg", "Caja x 30 tabletas", "Cardiovasculares", "#DC2626", 12500, 8000, 40, 10, 5, "2028-03-20", "Estante F-1"),
    ("AML5", "7701000000049", "Amlodipino 5 mg", "Caja x 30 tabletas", "Cardiovasculares", "#DC2626", 15000, 9500, 45, 10, 5, "2028-04-10", "Estante F-2"),
    ("ENL10", "7701000000050", "Enalapril 10 mg", "Caja x 30 tabletas", "Cardiovasculares", "#DC2626", 11000, 7000, 55, 12, 6, "2028-01-25", "Estante F-2"),
    ("HCT25", "7701000000051", "Hidroclorotiazida 25 mg", "Caja x 30 tabletas", "Cardiovasculares", "#DC2626", 8500, 5000, 60, 15, 8, "2028-07-30", "Estante F-3"),
    ("SIM40", "7701000000052", "Sinvastatina 40 mg", "Caja x 30 tabletas", "Cardiovasculares", "#DC2626", 25000, 16000, 30, 8, 4, "2028-02-15", "Estante F-3"),
    ("ASP100", "7701000000053", "Aspirina 100 mg", "Caja x 30 tabletas", "Cardiovasculares", "#DC2626", 7500, 4500, 90, 20, 10, "2028-09-10", "Estante F-4"),
    ("CAR10", "7701000000054", "Carvedilol 12.5 mg", "Caja x 30 tabletas", "Cardiovasculares", "#DC2626", 22000, 14500, 25, 6, 3, "2028-05-20", "Estante F-4"),

    # === DIABETES (6) ===
    ("MET850", "7701000000055", "Metformina 850 mg", "Caja x 30 tabletas", "Diabetes", "#059669", 8500, 5200, 80, 20, 10, "2028-08-15", "Estante G-1"),
    ("GLB5", "7701000000056", "Glibenclamida 5 mg", "Caja x 30 tabletas", "Diabetes", "#059669", 7000, 4200, 60, 15, 8, "2028-06-20", "Estante G-1"),
    ("GLP2", "7701000000057", "Glipizida 5 mg", "Caja x 30 tabletas", "Diabetes", "#059669", 12000, 7500, 40, 10, 5, "2028-04-10", "Estante G-2"),
    ("PIO15", "7701000000058", "Pioglitazona 30 mg", "Caja x 30 tabletas", "Diabetes", "#059669", 35000, 23000, 20, 5, 3, "2028-02-28", "Estante G-2"),
    ("ACR50", "7701000000059", "Acarbose 50 mg", "Caja x 30 tabletas", "Diabetes", "#059669", 28000, 18000, 25, 6, 3, "2028-01-15", "Estante G-3"),
    ("INS100", "7701000000060", "Insulina NPH 100 UI", "Frasco 10 ml", "Diabetes", "#059669", 45000, 30000, 15, 5, 2, "2027-09-30", "Nevera 1"),

    # === DERMATOLÓGICOS (8) ===
    ("HDC1", "7701000000061", "Hidrocremosa 1%", "Tubo 30 g", "Dermatológicos", "#DB2777", 6500, 3800, 70, 15, 8, "2027-11-20", "Estante H-1"),
    ("BMT005", "7701000000062", "Betametasona 0.05%", "Tubo 15 g", "Dermatológicos", "#DB2777", 12000, 7500, 40, 10, 5, "2027-08-15", "Estante H-1"),
    ("CLT2", "7701000000063", "Clotrimazol 2%", "Tubo 20 g", "Dermatológicos", "#DB2777", 9500, 6000, 45, 10, 5, "2028-03-10", "Estante H-2"),
    ("KNZ2", "7701000000064", "Ketoconazol 2%", "Tubo 15 g", "Dermatológicos", "#DB2777", 14000, 9000, 35, 8, 4, "2028-01-25", "Estante H-2"),
    ("MUP2", "7701000000065", "Mupirocina 2%", "Tubo 15 g", "Dermatológicos", "#DB2777", 22000, 14500, 25, 6, 3, "2027-10-30", "Estante H-3"),
    ("TERB1", "7701000000066", "Terbinafina 1%", "Tubo 15 g", "Dermatológicos", "#DB2777", 18000, 12000, 30, 8, 4, "2028-05-15", "Estante H-3"),
    ("ADA005", "7701000000067", "Adapaleno 0.05%", "Tubo 15 g", "Dermatológicos", "#DB2777", 28000, 18000, 20, 5, 3, "2027-12-20", "Estante H-4"),
    ("UREA20", "7701000000068", "Urea 20%", "Tubo 50 g", "Dermatológicos", "#DB2777", 11000, 7000, 38, 8, 4, "2028-07-10", "Estante H-4"),

    # === OFTÁLMICOS (6) ===
    ("CMZ05", "7701000000069", "Cloranfenicol colirio 0.5%", "Frasco 5 ml", "Oftálmicos", "#7C3AED", 8500, 5200, 50, 12, 6, "2027-09-15", "Estante I-1"),
    ("TOBRX", "7701000000070", "Tobramicina colirio 0.3%", "Frasco 5 ml", "Oftálmicos", "#7C3AED", 15000, 9500, 35, 8, 4, "2027-11-20", "Estante I-1"),
    ("DCLN", "7701000000071", "Diclofenaco colirio 0.1%", "Frasco 5 ml", "Oftálmicos", "#7C3AED", 12000, 7500, 40, 10, 5, "2028-02-10", "Estante I-2"),
    ("ARTL", "7701000000072", "Lágrimas Artificiales", "Frasco 15 ml", "Oftálmicos", "#7C3AED", 18000, 12000, 45, 10, 5, "2028-06-30", "Estante I-2"),
    ("CPLX", "7701000000073", "Ciprofloxacino colirio 0.3%", "Frasco 5 ml", "Oftálmicos", "#7C3AED", 11000, 7000, 30, 8, 4, "2027-10-15", "Estante I-3"),
    ("NAFZ", "7701000000074", "Nafazolina colirio", "Frasco 15 ml", "Oftálmicos", "#7C3AED", 7500, 4500, 55, 12, 6, "2028-04-20", "Estante I-3"),

    # === DISPOSITIVOS MÉDICOS (6) ===
    ("GLUC", "7701000000075", "Tiras Glucómetro x 50", "Caja", "Dispositivos Médicos", "#6366F1", 45000, 30000, 20, 5, 3, None, "Mostrador"),
    ("TENS", "7701000000076", "Tensiómetro Digital", "Unidad", "Dispositivos Médicos", "#6366F1", 85000, 55000, 10, 3, 2, None, "Mostrador"),
    ("OXIM", "7701000000077", "Oxímetro de Pulso", "Unidad", "Dispositivos Médicos", "#6366F1", 65000, 42000, 12, 4, 2, None, "Mostrador"),
    ("TERM", "7701000000078", "Termómetro Digital", "Unidad", "Dispositivos Médicos", "#6366F1", 25000, 16000, 30, 8, 4, None, "Mostrador"),
    ("VENDA", "7701000000079", "Vendas Elásticas x 3", "Paquete", "Dispositivos Médicos", "#6366F1", 8500, 5200, 50, 12, 6, None, "Estante J-1"),
    ("GASA", "7701000000080", "Gasas Estériles x 10", "Paquete", "Dispositivos Médicos", "#6366F1", 5500, 3200, 80, 20, 10, None, "Estante J-1"),

    # === CUIDADO PERSONAL (6) ===
    ("PROTX", "7701000000081", "Protector Solar FPS 50", "Frasco 60 ml", "Cuidado Personal", "#EC4899", 32000, 21000, 35, 8, 4, "2028-08-15", "Estante K-1"),
    ("REPH", "7701000000082", "Repelente Insectos", "Frasco 200 ml", "Cuidado Personal", "#EC4899", 18000, 12000, 50, 12, 6, "2028-05-20", "Estante K-1"),
    ("ALCOH", "7701000000083", "Alcohol Antiséptico 70%", "Botella 500 ml", "Cuidado Personal", "#EC4899", 7500, 4500, 60, 15, 8, None, "Estante K-2"),
    ("ISOT", "7701000000084", "Solución Salina 500 ml", "Frasco", "Cuidado Personal", "#EC4899", 5500, 3200, 70, 18, 9, None, "Estante K-2"),
    ("YODO", "7701000000085", "Yodo Povidona 120 ml", "Frasco", "Cuidado Personal", "#EC4899", 9000, 5500, 45, 10, 5, "2028-03-10", "Estante K-3"),
    ("AGUAOX", "7701000000086", "Agua Oxigenada 10 vol", "Frasco 250 ml", "Cuidado Personal", "#EC4899", 4500, 2800, 80, 20, 10, None, "Estante K-3"),

    # === ANTIFÚNGICOS (6) ===
    ("FLUC150", "7701000000087", "Fluconazol 150 mg", "Cápsula", "Antifúngicos", "#8B5CF6", 18000, 12000, 40, 10, 5, "2028-04-15", "Estante L-1"),
    ("NGST100", "7701000000088", "Nistatina 100,000 UI", "Caja x 20 tabletas", "Antifúngicos", "#8B5CF6", 9500, 6000, 50, 12, 6, "2028-01-20", "Estante L-1"),
    ("ITRC100", "7701000000089", "Itraconazol 100 mg", "Caja x 10 cápsulas", "Antifúngicos", "#8B5CF6", 28000, 18000, 25, 6, 3, "2027-11-30", "Estante L-2"),
    ("GRIS500", "7701000000090", "Griseofulvina 500 mg", "Caja x 20 tabletas", "Antifúngicos", "#8B5CF6", 14000, 9000, 30, 8, 4, "2027-09-25", "Estante L-2"),
    ("VRC200", "7701000000091", "Voriconazol 200 mg", "Caja x 10 tabletas", "Antifúngicos", "#8B5CF6", 85000, 55000, 10, 3, 2, "2027-12-15", "Estante L-3"),
    ("TRBN250", "7701000000092", "Terbinafina 250 mg", "Caja x 14 tabletas", "Antifúngicos", "#8B5CF6", 32000, 21000, 20, 5, 3, "2028-06-10", "Estante L-3"),

    # === ANTIPARASITARIOS (8) ===
    ("ALB400", "7701000000093", "Albendazol 400 mg", "Caja x 2 tabletas", "Antiparasitarios", "#F59E0B", 5500, 3200, 100, 25, 12, "2028-08-20", "Estante M-1"),
    ("MEB100", "7701000000094", "Mebendazol 100 mg", "Caja x 6 tabletas", "Antiparasitarios", "#F59E0B", 6500, 3800, 80, 20, 10, "2028-05-15", "Estante M-1"),
    ("IVM6", "7701000000095", "Ivermectina 6 mg", "Caja x 1 tableta", "Antiparasitarios", "#F59E0B", 12000, 7500, 70, 15, 8, "2028-02-28", "Estante M-2"),
    ("PNT50", "7701000000096", "Pirantel 250 mg", "Suspensión 30 ml", "Antiparasitarios", "#F59E0B", 8500, 5200, 45, 10, 5, "2027-11-10", "Estante M-2"),
    ("SECN", "7701000000097", "Secnidazol 2 g", "Sobre dosis única", "Antiparasitarios", "#F59E0B", 15000, 9500, 55, 12, 6, "2028-03-25", "Estante M-3"),
    ("TIN500", "7701000000098", "Tinidazol 500 mg", "Caja x 4 tabletas", "Antiparasitarios", "#F59E0B", 11000, 7000, 40, 10, 5, "2028-01-15", "Estante M-3"),
    ("PRAZ", "7701000000099", "Praziquantel 600 mg", "Caja x 3 tabletas", "Antiparasitarios", "#F59E0B", 22000, 14500, 20, 5, 3, "2027-10-20", "Estante M-4"),
    ("NIQZ", "7701000000100", "Nitazoxanida 500 mg", "Caja x 6 tabletas", "Antiparasitarios", "#F59E0B", 35000, 23000, 25, 6, 3, "2028-07-30", "Estante M-4"),
]

# Write data rows
for row_idx, product in enumerate(products, 2):
    for col_idx, value in enumerate(product, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name='Calibri', size=10)
        cell.alignment = Alignment(vertical='center')
        
        # Alternate row colors
        if row_idx % 2 == 0:
            cell.fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')

# Format number columns
price_cols = [7, 8]  # price, cost
for col in price_cols:
    for row in range(2, len(products) + 2):
        cell = ws.cell(row=row, column=col)
        cell.number_format = '#,##0'

# Column widths
col_widths = {
    'A': 12,  # code
    'B': 18,  # barcode
    'C': 35,  # name
    'D': 28,  # description
    'E': 22,  # category
    'F': 14,  # categoryColor
    'G': 12,  # price
    'H': 12,  # cost
    'I': 10,  # stock
    'J': 12,  # minStock
    'K': 14,  # criticalStock
    'L': 14,  # expiryDate
    'M': 15,  # location
}

for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Freeze header row
ws.freeze_panes = 'A2'

# Add filter
ws.auto_filter.ref = f'A1:M{len(products) + 1}'

# Save
output_path = r"c:\Users\UserMaster\Documents\MPointOfSale\productos_import.xlsx"
wb.save(output_path)
print(f"✅ Excel creado exitosamente: {output_path}")
print(f"📊 Total productos: {len(products)}")
print(f"📁 Columnas: {len(headers)}")
print(f"\nCategorías incluidas:")
cats = {}
for p in products:
    cats[p[4]] = cats.get(p[4], 0) + 1
for cat, count in sorted(cats.items(), key=lambda x: x[1], reverse=True):
    print(f"  • {cat}: {count}")
