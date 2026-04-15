import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ─── Sheet 1: Full sample (100 products) ─────────────────────────────────────
ws1 = wb.active
ws1.title = "Productos_Importar"

# ── Style constants ──
header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
data_font = Font(name='Calibri', size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
even_fill = PatternFill(start_color='F0F5FF', end_color='F0F5FF', fill_type='solid')

# ── Column headers (must match the import parser exactly) ──
headers = [
    'Código',
    'Código de Barras',
    'Nombre del Producto',
    'Categoría',
    'Precio Venta',
    'Costo',
    'Stock',
    'Stock Mínimo',
    'Stock Crítico',
    'Ubicación',
    'Fecha Vencimiento',
    'Tipo Unidad',
    'Factor Conversión',
    'Tasa Impuesto',
    'Cantidad Compra Sugerida',
    'Descripción',
]

row_num = 1
for col_num, header_text in enumerate(headers, 1):
    cell = ws1.cell(row=row_num, column=col_num, value=header_text)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# ── 100 realistic products across all 13 categories ──
products = [
    # ── Medicamentos (1-10) ──
    { 'code': 'MED001', 'barcode': '7701000000001', 'name': 'Acetaminofén 500mg', 'cat': 'Medicamentos', 'price': 5500, 'cost': 3200, 'stock': 100, 'min': 20, 'crit': 10, 'loc': 'Estante A1', 'exp': '2027-06-15', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 50, 'desc': 'Tabletas x 20 unidades' },
    { 'code': 'MED002', 'barcode': '7701000000002', 'name': 'Ibuprofeno 400mg', 'cat': 'Medicamentos', 'price': 8900, 'cost': 5100, 'stock': 80, 'min': 15, 'crit': 8, 'loc': 'Estante A1', 'exp': '2027-08-20', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 40, 'desc': 'Cápsulas x 10 unidades' },
    { 'code': 'MED003', 'barcode': '7701000000003', 'name': 'Amoxicilina 500mg', 'cat': 'Medicamentos', 'price': 15000, 'cost': 9500, 'stock': 60, 'min': 10, 'crit': 5, 'loc': 'Estante A2', 'exp': '2027-03-10', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 30, 'desc': 'Cápsulas x 21 unidades' },
    { 'code': 'MED004', 'barcode': '7701000000004', 'name': 'Loratadina 10mg', 'cat': 'Medicamentos', 'price': 7200, 'cost': 4300, 'stock': 90, 'min': 15, 'crit': 8, 'loc': 'Estante A2', 'exp': '2028-01-25', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 45, 'desc': 'Tabletas x 10 unidades' },
    { 'code': 'MED005', 'barcode': '7701000000005', 'name': 'Omeprazol 20mg', 'cat': 'Medicamentos', 'price': 12500, 'cost': 7800, 'stock': 70, 'min': 12, 'crit': 6, 'loc': 'Estante A3', 'exp': '2027-11-30', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 35, 'desc': 'Cápsulas x 30 unidades' },
    { 'code': 'MED006', 'barcode': '7701000000006', 'name': 'Diclofenaco 50mg', 'cat': 'Medicamentos', 'price': 6800, 'cost': 3900, 'stock': 85, 'min': 15, 'crit': 8, 'loc': 'Estante A3', 'exp': '2027-09-18', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 40, 'desc': 'Tabletas x 20 unidades' },
    { 'code': 'MED007', 'barcode': '7701000000007', 'name': 'Metformina 850mg', 'cat': 'Medicamentos', 'price': 9500, 'cost': 5800, 'stock': 75, 'min': 15, 'crit': 8, 'loc': 'Estante A4', 'exp': '2027-07-22', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 35, 'desc': 'Tabletas x 30 unidades' },
    { 'code': 'MED008', 'barcode': '7701000000008', 'name': 'Atenolol 50mg', 'cat': 'Medicamentos', 'price': 11000, 'cost': 6900, 'stock': 55, 'min': 10, 'crit': 5, 'loc': 'Estante A4', 'exp': '2027-12-05', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 25, 'desc': 'Tabletas x 30 unidades' },
    { 'code': 'MED009', 'barcode': '7701000000009', 'name': 'Losartán 50mg', 'cat': 'Medicamentos', 'price': 13500, 'cost': 8400, 'stock': 65, 'min': 12, 'crit': 6, 'loc': 'Estante A4', 'exp': '2027-10-14', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 30, 'desc': 'Tabletas x 30 unidades' },
    { 'code': 'MED010', 'barcode': '7701000000010', 'name': 'Salbutamol Inhalador', 'cat': 'Medicamentos', 'price': 18000, 'cost': 11500, 'stock': 30, 'min': 8, 'crit': 4, 'loc': 'Estante A5', 'exp': '2026-12-31', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 15, 'desc': 'Inhalador 100mcg x 200 dosis' },

    # ── Cuidado Personal (11-20) ──
    { 'code': 'CP001', 'barcode': '7702000000001', 'name': 'Protector Solar FPS50 60ml', 'cat': 'Cuidado Personal', 'price': 45000, 'cost': 28000, 'stock': 25, 'min': 5, 'crit': 3, 'loc': 'Estante B1', 'exp': '2027-04-01', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 15, 'desc': 'Protección UVA/UVB' },
    { 'code': 'CP002', 'barcode': '7702000000002', 'name': 'Crema Hidratante Facial 50g', 'cat': 'Cuidado Personal', 'price': 32000, 'cost': 19500, 'stock': 35, 'min': 8, 'crit': 4, 'loc': 'Estante B1', 'exp': '2027-06-15', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 20, 'desc': 'Piel sensible' },
    { 'code': 'CP003', 'barcode': '7702000000003', 'name': 'Shampoo Anticaída 400ml', 'cat': 'Cuidado Personal', 'price': 28000, 'cost': 17000, 'stock': 40, 'min': 10, 'crit': 5, 'loc': 'Estante B2', 'exp': '2028-01-01', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 20, 'desc': 'Con biotina y cafeína' },
    { 'code': 'CP004', 'barcode': '7702000000004', 'name': 'Jabón Antibacterial 200ml', 'cat': 'Cuidado Personal', 'price': 12000, 'cost': 7200, 'stock': 60, 'min': 15, 'crit': 8, 'loc': 'Estante B2', 'exp': '2028-03-15', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 30, 'desc': 'pH neutro' },
    { 'code': 'CP005', 'barcode': '7702000000005', 'name': 'Desodorante Roll-On 50ml', 'cat': 'Cuidado Personal', 'price': 15000, 'cost': 9000, 'stock': 50, 'min': 12, 'crit': 6, 'loc': 'Estante B2', 'exp': '2028-06-01', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 25, 'desc': 'Sin alcohol, 48h protección' },
    { 'code': 'CP006', 'barcode': '7702000000006', 'name': 'Pasta Dental Blanqueadora 100g', 'cat': 'Cuidado Personal', 'price': 9500, 'cost': 5800, 'stock': 70, 'min': 15, 'crit': 8, 'loc': 'Estante B3', 'exp': '2027-09-01', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 35, 'desc': 'Con flúor activo' },
    { 'code': 'CP007', 'barcode': '7702000000007', 'name': 'Enjuague Bucal 500ml', 'cat': 'Cuidado Personal', 'price': 14000, 'cost': 8500, 'stock': 45, 'min': 10, 'crit': 5, 'loc': 'Estante B3', 'exp': '2027-12-01', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 20, 'desc': 'Menta fresca, sin alcohol' },
    { 'code': 'CP008', 'barcode': '7702000000008', 'name': 'Crema para Manos 75ml', 'cat': 'Cuidado Personal', 'price': 18000, 'cost': 11000, 'stock': 55, 'min': 12, 'crit': 6, 'loc': 'Estante B3', 'exp': '2028-02-15', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 25, 'desc': 'Con aloe vera y vitamina E' },
    { 'code': 'CP009', 'barcode': '7702000000009', 'name': 'Protector Labial SPF15', 'cat': 'Cuidado Personal', 'price': 8500, 'cost': 5100, 'stock': 80, 'min': 15, 'crit': 8, 'loc': 'Mostrador', 'exp': '2028-04-01', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 40, 'desc': 'Con miel de abeja' },
    { 'code': 'CP010', 'barcode': '7702000000010', 'name': 'Toallas Húmedas x50', 'cat': 'Cuidado Personal', 'price': 7500, 'cost': 4500, 'stock': 90, 'min': 20, 'crit': 10, 'loc': 'Mostrador', 'exp': '2027-08-15', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 45, 'desc': 'Hipoalergénicas' },

    # ── Vitaminas y Suplementos (21-30) ──
    { 'code': 'VIT001', 'barcode': '7703000000001', 'name': 'Vitamina C 1000mg x30', 'cat': 'Vitaminas y Suplementos', 'price': 25000, 'cost': 15000, 'stock': 50, 'min': 10, 'crit': 5, 'loc': 'Estante C1', 'exp': '2028-03-01', 'unit': 'Unidad', 'factor': 1, 'tax': 0, 'qty': 25, 'desc': 'Tabletas efervescentes' },
    { 'code': 'VIT002', 'barcode': '7703000000002', 'name': 'Complejo B x60', 'cat': 'Vitaminas y Suplementos', 'price': 18000, 'cost': 10800, 'stock': 60, 'min': 12, 'crit': 6, 'loc': 'Estante C1', 'exp': '2028-05-15', 'unit': 'Unidad', 'factor': 1, 'tax': 0, 'qty': 30, 'desc': 'Tabletas recubiertas' },
    { 'code': 'VIT003', 'barcode': '7703000000003', 'name': 'Omega 3 EPA/DHA x60', 'cat': 'Vitaminas y Suplementos', 'price': 42000, 'cost': 26000, 'stock': 35, 'min': 8, 'crit': 4, 'loc': 'Estante C1', 'exp': '2028-01-20', 'unit': 'Unidad', 'factor': 1, 'tax': 0, 'qty': 18, 'desc': 'Cápsulas blandas' },
    { 'code': 'VIT004', 'barcode': '7703000000004', 'name': 'Vitamina D3 2000UI x30', 'cat': 'Vitaminas y Suplementos', 'price': 22000, 'cost': 13200, 'stock': 45, 'min': 10, 'crit': 5, 'loc': 'Estante C2', 'exp': '2028-06-10', 'unit': 'Unidad', 'factor': 1, 'tax': 0, 'qty': 22, 'desc': 'Cápsulas blandas' },
    { 'code': 'VIT005', 'barcode': '7703000000005', 'name': 'Hierro 30mg x30', 'cat': 'Vitaminas y Suplementos', 'price': 15000, 'cost': 9000, 'stock': 55, 'min': 12, 'crit': 6, 'loc': 'Estante C2', 'exp': '2028-04-25', 'unit': 'Unidad', 'factor': 1, 'tax': 0, 'qty': 28, 'desc': 'Tabletas' },
    { 'code': 'VIT006', 'barcode': '7703000000006', 'name': 'Calcio + Vit D x60', 'cat': 'Vitaminas y Suplementos', 'price': 28000, 'cost': 16800, 'stock': 40, 'min': 8, 'crit': 4, 'loc': 'Estante C2', 'exp': '2028-02-14', 'unit': 'Unidad', 'factor': 1, 'tax': 0, 'qty': 20, 'desc': 'Tabletas masticables' },
    { 'code': 'VIT007', 'barcode': '7703000000007', 'name': 'Magnesio 300mg x60', 'cat': 'Vitaminas y Suplementos', 'price': 32000, 'cost': 19200, 'stock': 38, 'min': 8, 'crit': 4, 'loc': 'Estante C3', 'exp': '2028-07-01', 'unit': 'Unidad', 'factor': 1, 'tax': 0, 'qty': 19, 'desc': 'Cápsulas' },
    { 'code': 'VIT008', 'barcode': '7703000000008', 'name': 'Zinc 15mg x60', 'cat': 'Vitaminas y Suplementos', 'price': 19000, 'cost': 11400, 'stock': 42, 'min': 10, 'crit': 5, 'loc': 'Estante C3', 'exp': '2028-08-15', 'unit': 'Unidad', 'factor': 1, 'tax': 0, 'qty': 21, 'desc': 'Tabletas' },
    { 'code': 'VIT009', 'barcode': '7703000000009', 'name': 'Multivitamínico x30', 'cat': 'Vitaminas y Suplementos', 'price': 35000, 'cost': 21000, 'stock': 30, 'min': 6, 'crit': 3, 'loc': 'Estante C3', 'exp': '2028-05-01', 'unit': 'Unidad', 'factor': 1, 'tax': 0, 'qty': 15, 'desc': 'Adultos mayores de 50' },
    { 'code': 'VIT010', 'barcode': '7703000000010', 'name': 'Colágeno Hidrolizado 300g', 'cat': 'Vitaminas y Suplementos', 'price': 48000, 'cost': 29000, 'stock': 20, 'min': 5, 'crit': 3, 'loc': 'Estante C3', 'exp': '2027-11-30', 'unit': 'Unidad', 'factor': 1, 'tax': 0, 'qty': 10, 'desc': 'Polvo soluble' },

    # ── Primeros Auxilios (31-40) ──
    { 'code': 'PA001', 'barcode': '7704000000001', 'name': 'Vendas Elásticas 5cm', 'cat': 'Primeros Auxilios', 'price': 8000, 'cost': 4800, 'stock': 45, 'min': 10, 'crit': 5, 'loc': 'Estante D1', 'exp': '2030-01-01', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 25, 'desc': 'Pack x3 unidades' },
    { 'code': 'PA002', 'barcode': '7704000000002', 'name': 'Curitas Adhesivas x20', 'cat': 'Primeros Auxilios', 'price': 5500, 'cost': 3300, 'stock': 80, 'min': 15, 'crit': 8, 'loc': 'Estante D1', 'exp': '2030-06-01', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 40, 'desc': 'Tela resistente al agua' },
    { 'code': 'PA003', 'barcode': '7704000000003', 'name': 'Gasas Estériles 10x10cm x10', 'cat': 'Primeros Auxilios', 'price': 7000, 'cost': 4200, 'stock': 60, 'min': 12, 'crit': 6, 'loc': 'Estante D1', 'exp': '2030-03-15', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 30, 'desc': 'No adherentes' },
    { 'code': 'PA004', 'barcode': '7704000000004', 'name': 'Alcohol Antiséptico 250ml', 'cat': 'Primeros Auxilios', 'price': 6500, 'cost': 3900, 'stock': 70, 'min': 15, 'crit': 8, 'loc': 'Estante D2', 'exp': '2028-12-01', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 35, 'desc': '70% v/v' },
    { 'code': 'PA005', 'barcode': '7704000000005', 'name': 'Agua Oxigenada 120ml', 'cat': 'Primeros Auxilios', 'price': 4500, 'cost': 2700, 'stock': 75, 'min': 15, 'crit': 8, 'loc': 'Estante D2', 'exp': '2028-10-15', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 35, 'desc': '3% volumen' },
    { 'code': 'PA006', 'barcode': '7704000000006', 'name': 'Yodo Povidona 120ml', 'cat': 'Primeros Auxilios', 'price': 8500, 'cost': 5100, 'stock': 50, 'min': 10, 'crit': 5, 'loc': 'Estante D2', 'exp': '2028-08-01', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 25, 'desc': 'Solución tópica 10%' },
    { 'code': 'PA007', 'barcode': '7704000000007', 'name': 'Algodón Hidrófilo 100g', 'cat': 'Primeros Auxilios', 'price': 5000, 'cost': 3000, 'stock': 65, 'min': 12, 'crit': 6, 'loc': 'Estante D3', 'exp': '2030-01-01', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 30, 'desc': 'No estéril' },
    { 'code': 'PA008', 'barcode': '7704000000008', 'name': 'Cinta Adhesiva Microporosa', 'cat': 'Primeros Auxilios', 'price': 9000, 'cost': 5400, 'stock': 40, 'min': 8, 'crit': 4, 'loc': 'Estante D3', 'exp': '2030-06-01', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 20, 'desc': '2.5cm x 5m' },
    { 'code': 'PA009', 'barcode': '7704000000009', 'name': 'Guantes Desechables x10', 'cat': 'Primeros Auxilios', 'price': 12000, 'cost': 7200, 'stock': 55, 'min': 10, 'crit': 5, 'loc': 'Estante D3', 'exp': '2030-01-01', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 25, 'desc': 'Nitrilo, talla M' },
    { 'code': 'PA010', 'barcode': '7704000000010', 'name': 'Termómetro Digital', 'cat': 'Primeros Auxilios', 'price': 22000, 'cost': 13500, 'stock': 15, 'min': 5, 'crit': 3, 'loc': 'Mostrador', 'exp': None, 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 8, 'desc': 'Punta flexible' },

    # ── Dispositivos Médicos (41-50) ──
    { 'code': 'DM001', 'barcode': '7705000000001', 'name': 'Tensiómetro Digital Brazo', 'cat': 'Dispositivos Médicos', 'price': 85000, 'cost': 55000, 'stock': 10, 'min': 3, 'crit': 2, 'loc': 'Mostrador', 'exp': None, 'unit': 'Unidad', 'factor': 1, 'tax': 0, 'qty': 5, 'desc': 'Memoria 120 lecturas' },
    { 'code': 'DM002', 'barcode': '7705000000002', 'name': 'Oxímetro de Pulso', 'cat': 'Dispositivos Médicos', 'price': 65000, 'cost': 42000, 'stock': 12, 'min': 4, 'crit': 2, 'loc': 'Mostrador', 'exp': None, 'unit': 'Unidad', 'factor': 1, 'tax': 0, 'qty': 6, 'desc': 'Pantalla OLED' },
    { 'code': 'DM003', 'barcode': '7705000000003', 'name': 'Glucómetro Kit Completo', 'cat': 'Dispositivos Médicos', 'price': 75000, 'cost': 48000, 'stock': 8, 'min': 3, 'crit': 2, 'loc': 'Mostrador', 'exp': None, 'unit': 'Unidad', 'factor': 1, 'tax': 0, 'qty': 4, 'desc': 'Incluye 10 tiras' },
    { 'code': 'DM004', 'barcode': '7705000000004', 'name': 'Tiras Reactivas Glucosa x50', 'cat': 'Dispositivos Médicos', 'price': 45000, 'cost': 28000, 'stock': 20, 'min': 5, 'crit': 3, 'loc': 'Estante E1', 'exp': '2027-06-01', 'unit': 'Unidad', 'factor': 1, 'tax': 0, 'qty': 10, 'desc': 'Compatible con glucómetro DM003' },
    { 'code': 'DM005', 'barcode': '7705000000005', 'name': 'Nebulizador Eléctrico', 'cat': 'Dispositivos Médicos', 'price': 120000, 'cost': 78000, 'stock': 5, 'min': 2, 'crit': 1, 'loc': 'Estante E1', 'exp': None, 'unit': 'Unidad', 'factor': 1, 'tax': 0, 'qty': 3, 'desc': 'Portátil, bajo ruido' },
    { 'code': 'DM006', 'barcode': '7705000000006', 'name': 'Mascarilla N95 x5', 'cat': 'Dispositivos Médicos', 'price': 15000, 'cost': 9000, 'stock': 40, 'min': 10, 'crit': 5, 'loc': 'Estante E1', 'exp': '2028-12-01', 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 20, 'desc': 'Con válvula de exhalación' },
    { 'code': 'DM007', 'barcode': '7705000000007', 'name': 'Bolsa Agua Caliente 1L', 'cat': 'Dispositivos Médicos', 'price': 18000, 'cost': 11000, 'stock': 20, 'min': 5, 'crit': 3, 'loc': 'Estante E2', 'exp': None, 'unit': 'Unidad', 'factor': 1, 'tax': 0.19, 'qty': 10, 'desc': 'PVC resistente' },
    { 'code': 'DM008', 'barcode': '7705000000008', 'name': 'Bolsa Gel Frío